#This is my steel calculator script

#This section of code will create the graphical user interface (GUI)

#load in steel functions which perform operations

#Choose sectiontype and size
from decimal import Decimal
import PySimpleGUI as sg
import steel_functions as st
import Bolt_checks as bc
import math
import sys
import os
import csv
script_dir = os.path.dirname( __file__ )
count = 0
for x in script_dir:
    count += 1
    if x == '\\':
        count1 = count
x = len(script_dir)-count1
script_dir = script_dir[:-x]
mymodule_dir = os.path.join( script_dir, 'Timber' )
sys.path.append( mymodule_dir )
mymodule_dir1 = os.path.join( script_dir, 'Concrete' )
sys.path.append( mymodule_dir1 )
mymodule_dir1 = os.path.join( script_dir, 'Deflection' )
sys.path.append( mymodule_dir1 )
from Timber import beam_moment, beam_shear,print_calcs,deflection_check
import Concrete as Concrete
import Deflection as Deflection
import Retaining_wall as Retaining_wall_script
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('TkAgg')
from matplotlib.ticker import NullFormatter
import numpy as np

def menu(variables):
    sg.theme('GrayGrayGray')
    layout1 = [
        [sg.Button('Steel Calculator', key = 'Steel_calculator')],
        [sg.Button('Bolt group actions', key = 'Bolts')],
        [sg.Button('Development Length', key = 'development')],
        [sg.Button('Timber Calculator', key = 'Timber_calculator')],
        [sg.Button('Concrete Beam Calculator', key = 'Concrete_calculator')],
        [sg.Button('Retaining Wall Calculator', key = 'Retaining_wall')],
        [sg.Button('Soldier Pile Retaining wall', key = 'Soldier pile')],
        [sg.Button('Laterally Loaded Pile', key='Laterally loaded pile')],
        [sg.Button('Pad Footing Calculator', key ='Pad_footing')],
        [sg.Button('Wind Calculator', key='Wind_calculator')]
    ]
    window1 = sg.Window('Menu Select', layout1)

    while True:
        event, values1 = window1.read()
        if event == 'Steel_calculator':
            window1.close()
            steel_calculator()
            break
        elif event == 'Bolts':
            window1.close()
            bolts()
            break
        elif event == 'development':
            window1.close()
            development()
            break
        elif event == 'Timber_calculator':
            window1.close()
            Timber()
            break
        elif event == 'Concrete_calculator':
            window1.close()
            Concrete_Beam(concrete_beam)
            break
        elif event == 'Wind_calculator':
            window1.close()
            Wind()
        elif event == 'Retaining_wall':
            window1.close()
            Retaining_wall()
        elif event == 'Soldier pile':
            window1.close()
            Soldier_pile()
        elif event == 'Laterally loaded pile':
            window1.close()
            Lateral_pile()
        elif event == 'Pad_footing':
            window1.close()
            Pad_footing()
        elif event == sg.WIN_CLOSED:
            break
    window1.close()

def variable_write(values,Name):
    #if values['Member Type'] == 'Beam':
        with open(Name +'.txt','w',newline='') as csv_file:
            data = [[str(i),values[i]] for i in values]
            print(data)
            writer = csv.writer(csv_file)
            writer.writerows(data)


def variable(Name):
    Dictionary = {}
    with open(str(Name) + '.txt') as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',')
        for row in csv_reader:
            Dictionary[row[0]] = row[1]
    return Dictionary

def KeyCheck(Dictionary, string):
    try:
        output = Dictionary[string]
    except KeyError:
        output = "BLANK"
    return output

def isfloat(num):
    try:
        float(num)
        return True
    except ValueError:
        return False


def Lateral_pile():
    layout = [
        [sg.Column([
            [sg.Text('Soil Parameters')],
            [sg.Text('Friction angle:')],
            [sg.Text('Density of soil:')],
            [sg.Text('Cohesion:')],
            [sg.Text('Angle of Backfill:')],
            [sg.Text('Ka')],
            [sg.Text('Kp')],
            [sg.Text('Retained Height')],
            [sg.Text('Pile Diameter')],
            [sg.Text('Pile spacing')],
            [sg.Text('Surcharge')],

            [sg.Text('Height of sleeper:')],
            [sg.Text('Depth to Water Table:')]


        ]),sg.Column([
            [sg.Text()],
            [sg.Input(default_text=26,size=(5,1),key='Friction_angle')],
            [sg.Input(default_text=20,size=(5,1),key='Density_soil')],
            [sg.Input(default_text=5,size=(5,1),key='cohesion')],
            [sg.Input(default_text=15,size=(5,1),key='beta')],
            [sg.Input(default_text=0.4,size=(5,1),key='Ka')],
            [sg.Input(default_text=2.9, size=(5, 1), key='Kp')],
            [sg.Input(default_text=1.5,size=(5,1),key='H')],
            [sg.Input(default_text=450,size=(5,1),key='Dia')],
            [sg.Input(default_text=2.25,size=(5,1),key ='Spacing')],
            [sg.Input(default_text=5,size=(5,1),key='surcharge')],


            [sg.Input(default_text=400, size=(5, 1), key='SleeperH')],
            [sg.Input(default_text=0,size=(5,1),key='Water_table')]
        ]),sg.Column([
            [sg.Text()],
            [sg.Text('Degrees')],
            [sg.Text('KN/m3')],
            [sg.Text('KPa')],
            [sg.Text('Degrees')],
            [sg.Checkbox(default=True,text='Override',key='Ka_over')],
            [sg.Checkbox(default=True,text='Override',key='Kp_over')],

            [sg.Text('m')],
            [sg.Text('mm')],
            [sg.Text('mm')],
            [sg.Text('KPa')],

            [sg.Text('mm')],
            [sg.Text('m')],

        ]),sg.Column(
            blank(5) + [[sg.Text(key='Ka1')],[sg.Text(key='Kp1')]] + blank(5) +[[sg.Checkbox(default=True, text='Water Table', key='Water')]]
        )

    ],
    [sg.Text('Results:')],
    [sg.Column([
        [sg.Text('d:')],
        [sg.Text('D:')],
        [sg.Text('Total Embedment, E:')],
        [sg.Text('Soil Force, Pa:')],
        [sg.Text('Surcharge Force, Pw:')],
        [sg.Text('Max Moment:')],
        [sg.Text('Effective pile width factor:')],

        [sg.Text('Max moment on sleeper:')],
        [sg.Text('Max shear on sleeper:')]
    ]),sg.Column([
        [sg.Text(key='d')],
        [sg.Text(key='D')],
        [sg.Text(key='E')],
        [sg.Text(key='Pa')],
        [sg.Text(key='Pw')],
        [sg.Text(key='Mmax')],
        [sg.Text(key='f')],

        [sg.Text(key='M')],
        [sg.Text(key='V')]
    ]),sg.Column([
        [sg.Text('m')],
        [sg.Text('m')],
        [sg.Text('m')],
        [sg.Text('KN')],
        [sg.Text('KN')],
        [sg.Text('KNm')],
        [sg.Text()],

        [sg.Text('KNm')],
        [sg.Text('KN')]
    ])],
        [sg.Text()],
        [sg.Button('Calculate', key='Calculate'), sg.Button('Back', key='back')],
        [sg.Button('Print calculations', key='print_calcs')],
        [sg.Text('Type Job Name:'), sg.Input(default_text='Soldier Pile', key='job_name')],
        [sg.Text('Choose destination:'),
         sg.Input(key='print_location', default_text=r'C:\Users\tduffett\PycharmProjects\pythonProject1'),
         sg.FolderBrowse()],
        [sg.Button('Back', key='back')],
    ]

    window = sg.Window('Soldier Pile Retaining wall',layout)

    while True:
        event, values = window.read()
        if event == sg.WIN_CLOSED:
            break
        elif event == 'Calculate':
            for i in values:
                if isfloat(values[i]) == True:
                    values[i] = float(values[i])
            if values['Ka_over'] == False:
                values['Ka'] = (math.cos(values['beta']/180*math.pi) - math.sqrt(math.cos(values['beta']/180*math.pi)**2 - math.cos(values['Friction_angle']/180*math.pi)**2))/(math.cos(values['beta']/180*math.pi) + math.sqrt(math.cos(values['beta']/180*math.pi)**2 - math.cos(values['Friction_angle']/180*math.pi)**2))
                window['Ka1'].update(str(round(values['Ka'],2)))
            if values['Kp_over'] == False:
                values['Kp'] = math.tan((45 + values['Friction_angle']/2)/180*math.pi)**2
                window['Kp1'].update(str(round(values['Kp'], 2)))
            Result = Retaining_wall_script.Soldier(values)
            for i in Result:
                window[i].update(str(round(Result[i],3)))



def Soldier_pile():
    layout = [
        [sg.Column([
            [sg.Text('Soil Parameters')],
            [sg.Text('Friction angle:')],
            [sg.Text('Density of soil:')],
            [sg.Text('Cohesion:')],
            [sg.Text('Angle of Backfill:')],
            [sg.Text('Ka')],
            [sg.Text('Kp')],
            [sg.Text('Retained Height')],
            [sg.Text('Pile Diameter')],
            [sg.Text('Pile spacing')],
            [sg.Text('Surcharge')],

            [sg.Text('Height of sleeper:')],
            [sg.Text('Depth to Water Table:')]


        ]),sg.Column([
            [sg.Text()],
            [sg.Input(default_text=26,size=(5,1),key='Friction_angle')],
            [sg.Input(default_text=20,size=(5,1),key='Density_soil')],
            [sg.Input(default_text=5,size=(5,1),key='cohesion')],
            [sg.Input(default_text=15,size=(5,1),key='beta')],
            [sg.Input(default_text=0.4,size=(5,1),key='Ka')],
            [sg.Input(default_text=2.9, size=(5, 1), key='Kp')],
            [sg.Input(default_text=1.5,size=(5,1),key='H')],
            [sg.Input(default_text=450,size=(5,1),key='Dia')],
            [sg.Input(default_text=2.25,size=(5,1),key ='Spacing')],
            [sg.Input(default_text=5,size=(5,1),key='surcharge')],


            [sg.Input(default_text=400, size=(5, 1), key='SleeperH')],
            [sg.Input(default_text=0,size=(5,1),key='Water_table')]
        ]),sg.Column([
            [sg.Text()],
            [sg.Text('Degrees')],
            [sg.Text('KN/m3')],
            [sg.Text('KPa')],
            [sg.Text('Degrees')],
            [sg.Checkbox(default=True,text='Override',key='Ka_over')],
            [sg.Checkbox(default=True,text='Override',key='Kp_over')],

            [sg.Text('m')],
            [sg.Text('mm')],
            [sg.Text('mm')],
            [sg.Text('KPa')],

            [sg.Text('mm')],
            [sg.Text('m')],

        ]),sg.Column(
            blank(5) + [[sg.Text(key='Ka1')],[sg.Text(key='Kp1')]] + blank(5) +[[sg.Checkbox(default=True, text='Water Table', key='Water')]]
        )

    ],
    [sg.Text('Results:')],
    [sg.Column([
        [sg.Text('d:')],
        [sg.Text('D:')],
        [sg.Text('Total Embedment, E:')],
        [sg.Text('Soil Force, Pa:')],
        [sg.Text('Surcharge Force, Pw:')],
        [sg.Text('Max Moment:')],
        [sg.Text('Effective pile width factor:')],

        [sg.Text('Max moment on sleeper:')],
        [sg.Text('Max shear on sleeper:')]
    ]),sg.Column([
        [sg.Text(key='d')],
        [sg.Text(key='D')],
        [sg.Text(key='E')],
        [sg.Text(key='Pa')],
        [sg.Text(key='Pw')],
        [sg.Text(key='Mmax')],
        [sg.Text(key='f')],

        [sg.Text(key='M')],
        [sg.Text(key='V')]
    ]),sg.Column([
        [sg.Text('m')],
        [sg.Text('m')],
        [sg.Text('m')],
        [sg.Text('KN')],
        [sg.Text('KN')],
        [sg.Text('KNm')],
        [sg.Text()],

        [sg.Text('KNm')],
        [sg.Text('KN')]
    ])],
        [sg.Text()],
        [sg.Button('Calculate', key='Calculate'), sg.Button('Back', key='back')],
        [sg.Button('Print calculations', key='print_calcs')],
        [sg.Text('Type Job Name:'), sg.Input(default_text='Soldier Pile', key='job_name')],
        [sg.Text('Choose destination:'),
         sg.Input(key='print_location', default_text=r'C:\Users\tduffett\PycharmProjects\pythonProject1'),
         sg.FolderBrowse()],
        [sg.Button('Back', key='back')],
    ]

    window = sg.Window('Soldier Pile Retaining wall',layout)

    while True:
        event, values = window.read()
        if event == sg.WIN_CLOSED:
            break
        elif event == 'Calculate':
            for i in values:
                if isfloat(values[i]) == True:
                    values[i] = float(values[i])
            if values['Ka_over'] == False:
                values['Ka'] = (math.cos(values['beta']/180*math.pi) - math.sqrt(math.cos(values['beta']/180*math.pi)**2 - math.cos(values['Friction_angle']/180*math.pi)**2))/(math.cos(values['beta']/180*math.pi) + math.sqrt(math.cos(values['beta']/180*math.pi)**2 - math.cos(values['Friction_angle']/180*math.pi)**2))
                window['Ka1'].update(str(round(values['Ka'],2)))
            if values['Kp_over'] == False:
                values['Kp'] = math.tan((45 + values['Friction_angle']/2)/180*math.pi)**2
                window['Kp1'].update(str(round(values['Kp'], 2)))
            Result = Retaining_wall_script.Soldier(values)
            for i in Result:
                window[i].update(str(round(Result[i],3)))




def Pad_footing_write(values):
    with open('Pf.txt','w',newline='') as csv_file:
        data = [

        ]

        writer = csv.writer(csv_file)
        writer.writerows(data)

def Pad_footing_read():
    Pf = {}
    with open('Pf.txt') as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',')
        for row in csv_reader:
            print(row)
            if row[0] == 'H':
                Pf['H'] = round(float(row[1]))
    return Pf

def Pad_footing():
    Pf = Pad_footing_read()

    layout = [
        [sg.Text('Pad Footing Calculator')],
        []

    ]

def blank(n):
    layout =[[]]
    for i in range(n):
        layout += [[sg.Text('')]]
    return layout
def Retaining_wall():
    #Retaining_wall = Retaining_wall_read()
    Retaining_wall = variable('Retaining_wall')
    list1 = ['Unrestrained','Minor','Moderate','Strong']
    if Retaining_wall['Masonry'] == 'False':
        Retaining_wall['Masonry'] = False
    layout = [
        [sg.Column([
        [sg.Text('Retaining Wall Calculator')],
        [sg.Text('Pure Cantilever'),sg.Checkbox('Masonry wall',key='Masonry',enable_events=True,default=Retaining_wall['Masonry'])],
        [sg.Text('Dimensions:')],
        [sg.Text('Wall Height, H:       ')],
        [sg.Text('Footing Depth, D:     ')],
        [sg.Text('Length of Toe, Ltoe:  ')],
        [sg.Text('Length of Heel, Lheel:')],


        [sg.Text('Soil Properties:')],
        [sg.Text('Angle of internal friction, \u03A6')],
        [sg.Text('Soil cohesion, c\'                ')],
        [sg.Text('Density of soil, \u03B3')],
        [sg.Text('Allowable bearing capacity:')],
        [sg.Text('Passive Soil coefficient:')],
        [sg.Text('Active Soil coefficient:')],
        [sg.Text('Loads:')],
        [sg.Text('Surcharge')],
        [sg.Text('Additional Moment')],
        [sg.Text('Additional Load ontop of Wall')],
        [sg.Checkbox('Restrained at Top',default=KeyCheck(Retaining_wall,'Top_restraint'),key='Top_restraint',enable_events=True)],
        [sg.Text('Results:')],
        [sg.Text('Bearing Pressure (SLS):')],
        [sg.Text('Overturning (ULS):')],
        [sg.Text('Sliding:')],
        [sg.Text('Wall vertical steel:')],
        [sg.Text('Wall horizontal steel:')],
        [sg.Text('Toe reinforcement:')],
        [sg.Text('Toe shrinkage steel:')],
        [sg.Text('Heel reinforcement:')],
        [sg.Text('Heel shrinkage steel:')]
        ]),sg.Column([
            [sg.Text('')],
            [sg.Text('')],
            [sg.Text('')],
            [sg.Input(default_text=KeyCheck(Retaining_wall,'H'), size=(5, 1), key='H',enable_events=True)],
            [sg.Input(default_text=KeyCheck(Retaining_wall,'D'), size=(5, 1), key='D',enable_events=True)],
            [sg.Input(default_text=KeyCheck(Retaining_wall,'Ltoe'), size=(5, 1), key='Ltoe',enable_events=True)],
            [sg.Input(default_text=KeyCheck(Retaining_wall,'Lheel'), size=(5, 1), key='Lheel',enable_events=True)],
            [sg.Text('')],
            [sg.Input(default_text=KeyCheck(Retaining_wall,'Internal_friction'), size=(5, 1), key='Internal_friction',enable_events=True)],
            [sg.Input(default_text=KeyCheck(Retaining_wall,'c'), size=(5, 1), key='c',enable_events=True)],
            [sg.Input(default_text=KeyCheck(Retaining_wall,'Soil_density'), size=(5, 1), key='Soil_density',enable_events=True)],
            [sg.Input(default_text=KeyCheck(Retaining_wall,'Allowable_bearing'),size=(5,1),key = 'Allowable_bearing',enable_events=True)],
            [sg.Input(default_text=KeyCheck(Retaining_wall,'Kp'), size=(5, 1), key='Kp',enable_events=True)],
            [sg.Input(default_text=KeyCheck(Retaining_wall,'Ka'), size=(5, 1), key='Ka',enable_events=True)],
            [sg.Text('')],
            [sg.Input(default_text=KeyCheck(Retaining_wall,'Surcharge'), size=(5, 1), key='Surcharge',enable_events=True)],
            [sg.Input(default_text=KeyCheck(Retaining_wall,'Additional_moment'), size=(5, 1), key='Additional_moment',enable_events=True)],
            [sg.Input(default_text=KeyCheck(Retaining_wall,'Additional_load'), size=(5, 1), key='Additional_load',
                      enable_events=True)],
            [sg.Input(default_text=KeyCheck(Retaining_wall,'Restraint_height'),size=(5,1),key='Restraint_height',enable_events=True)],
            [sg.Text('')],

            [sg.Text('',key='bearing_check')],
            [sg.Text('',key='overturning_check')],
            [sg.Text('',key='sliding_check')],
            [sg.Text('',key='wall_vertical_steel')],
            [sg.Text('',key='wall_horizontal_steel')],
            [sg.Text('',key='toe_horizontal_steel')],
            [sg.Text('',key='toe_OoP')],
            [sg.Text('',key='heel_horizontal_steel')],
            [sg.Text('',key='heel_OoP')]
        ]),sg.Column([
            [sg.Text('')],
            [sg.Text('')],
            [sg.Text('')],
            [sg.Text('mm')],
            [sg.Text('mm')],
            [sg.Text('mm')],
            [sg.Text('mm')],
            [sg.Text('')],
            [sg.Text('Degrees')],
            [sg.Text('KPa')],
            [sg.Text('KN/m3')],
            [sg.Text('KPa')],
            [sg.Checkbox(text='Override',key='Kp_over')],
            [sg.Checkbox(text='Override',key='Ka_over')],
            [sg.Text('')],
            [sg.Text('KPa')],
            [sg.Text('KNm')],
            [sg.Text('KN')],
            [sg.Text('Restraint Height mm')],
            [sg.Text('')],
            [sg.Text('KPa')],
            [sg.Text('KNm')],
            [sg.Text('KN')],
            [sg.Text('mm2/m')],
            [sg.Text('mm2/m')],
            [sg.Text('mm2/m')],
            [sg.Text('mm2/m')],
            [sg.Text('mm2/m')],
            [sg.Text('mm2/m')],
        ]),sg.Column(blank(12)+[
            [sg.Text(key='Kp1')],
            [sg.Text(key='Ka1')]
        ] + blank(6)+[

            [sg.Text('OK/NG',key='bearing_OK/NG')],
            [sg.Text('OK/NG',key='overturning_OK/NG')],
            [sg.Text('OK/NG',key='sliding_OK/NG')],
            [sg.Text('minimum')],
            [sg.Text('minimum')],
            [sg.Text('')],
            [sg.Text('')],
            [sg.Text('')],
            [sg.Text('')]
        ]),sg.Column(blank(2)+[
            [sg.Text('Wall thickness, tw:')],
            [sg.Text('Shear Key Depth:')],
            [sg.Text('Shear Key Width:')],
            [sg.Text('Shear Key Distance from Toe:')],
            [sg.Text('Angle of Backfill soil:')],
            [sg.Text('Overburden Soil:')]
        ]+blank(1)+[
            [sg.Text('Concrete Properties')],
            [sg.Text('f\'c')],
            [sg.Text('Concrete Density')],
            [sg.Text('Wall cover')],
            [sg.Text('Footing cover:')],
            [sg.Text('Degree of Crack control')],
            [sg.Text('')],
            [sg.Text('Masonry Properties')],
            [sg.Text('fuc')],
            [sg.Text('Bedding thickness:')],
            [sg.Text('Masonry Block Size:')],
            [sg.Text('Masonry Material:')],
            [sg.Text('Depth from extreme comp. to steel, d:')]

        ]+blank(7)),sg.Column(blank(2) + [
            [sg.Input(default_text=KeyCheck(Retaining_wall,'tw'), size=(5, 1), key='tw', enable_events=True)],
            [sg.Input(default_text=KeyCheck(Retaining_wall,'SD'), size=(5, 1), key='SD', enable_events=True)],
            [sg.Input(default_text=KeyCheck(Retaining_wall,'SW'), size=(5, 1), key='SW', enable_events=True)],
            [sg.Input(default_text=KeyCheck(Retaining_wall,'SKD'), size=(5, 1), key='SKD', enable_events=True)],
            [sg.Input(default_text=KeyCheck(Retaining_wall,'beta'), size=(5, 1), key='beta', enable_events=True)],
            [sg.Input(default_text=KeyCheck(Retaining_wall,'Overburden'), size = (5,1), key = 'Overburden', enable_events=True)]
        ]+blank(2)+[

            [sg.Input(default_text=KeyCheck(Retaining_wall,'fc'), size=(5, 1), key='fc', enable_events=True)],
            [sg.Input(default_text=KeyCheck(Retaining_wall,'DC'), size=(5, 1), key='DC', enable_events=True)],
            [sg.Input(default_text=KeyCheck(Retaining_wall,'wall_cover'), size=(5, 1), key='wall_cover', enable_events=True)],
            [sg.Input(default_text=KeyCheck(Retaining_wall,'footing_cover'), size=(5, 1), key='footing_cover', enable_events=True)],
            [sg.Combo(list1, default_value=list1[0],
                      size=(12, 1), key='crack', enable_events=True)],
            [sg.Text('')],
            [sg.Text('')],
            [sg.Input(default_text=KeyCheck(Retaining_wall,'fuc'), size=(5, 1), key='fuc', enable_events=True)],
            [sg.Input(default_text=KeyCheck(Retaining_wall,'BT'), size=(5, 1), key='BT', enable_events=True)],
            [sg.Combo(['10 series','15 series','20 series','25 series','30 series'], default_value='20 series',
                      size=(12, 1), key='Masonry_size', enable_events=True)],
            [sg.Combo(['Clay','Concrete'], default_value='Clay',
                      size=(12, 1), key='Masonry_type', enable_events=True)],
            [sg.Input(default_text=120, size=(5,1),key='d',enable_events=True)]
        ]+blank(7)),sg.Column(blank(2)+[
            [sg.Text('mm')],
            [sg.Text('mm')],
            [sg.Text('mm')],
            [sg.Text('mm')],
            [sg.Text('Degrees')],
            [sg.Text('mm')]

        ] + blank(2)+[
            [sg.Text('MPa')],
            [sg.Text('KN/m3')],
            [sg.Text('mm')],
            [sg.Text('mm')]
        ]+blank(14)),sg.Column(blank(27))],
        [sg.Button('Calculate',key='Calculate'),sg.Button('Back',key = 'back')],
        [sg.Button('Print calculations', key='print_calcs')],
        [sg.Text('Type Job Name:'), sg.Input(default_text='Retaining Walls', key='job_name')],
        [sg.Text('Choose destination:'),
         sg.Input(key='print_location', default_text=r'C:\Users\tduffett\PycharmProjects\pythonProject1'),
         sg.FolderBrowse()],
        [sg.Button('Back', key='back')],

    ]
    window = sg.Window('Retaining Wall Calculator', layout,resizable=True).finalize()
    window.Maximize()

    while True:
        event, values = window.read()
        for i in values:
            if event == i:
                variable_write(values,'Retaining_wall')
        if event == sg.WIN_CLOSED:
            break
        elif event == 'print_calcs':
            Retaining_wall_script.print_calcs(values)
        #elif event == 'H' or event == 'Overburden' or event == 'Additional_load' or event == 'SKD' or event == 'fuc' or event == 'BT' or event == 'D' or event == 'Masonry' or event == 'crack_masonry' or event == 'list1' or event == 'footing_cover' or event == 'crack' or event == 'cover' or event == 'Ka' or event == 'Kp' or event == 'Additional_moment' or event == 'Surcharge' or event == 'Ltoe' or event == 'beta' or event == 'Lheel' or event == 'tw' or event == 'Internal_friction' or event == 'c' or event == 'Soil_density' or event == 'SD' or event == 'SW' or event == 'Allowable_bearing' or event == 'fc' or event == 'DC':

            #Retaining_wall_write(values)
        elif event == 'Calculate':
            for i in values:
                if isfloat(values[i]) == True:
                    values[i] = float(values[i])
            if values['Ka_over'] == False:
                values['Ka'] = (math.cos(values['beta']/180*math.pi) - math.sqrt(math.cos(values['beta']/180*math.pi)**2 - math.cos(values['Internal_friction']/180*math.pi)**2))/(math.cos(values['beta']/180*math.pi) + math.sqrt(math.cos(values['beta']/180*math.pi)**2 - math.cos(values['Internal_friction']/180*math.pi)**2))
                window['Ka1'].update(str(round(values['Ka'],2)))
            if values['Kp_over'] == False:
                values['Kp'] = math.tan((45 + values['Internal_friction']/2)/180*math.pi)**2
                window['Kp1'].update(str(round(values['Kp'], 2)))


            h = math.tan(float(values['beta'])/180*math.pi)*float(values['Lheel'])
            Dimensions = {'SW':float(values['SW']),'SD':float(values['SD']),'SKD':float(values['SKD']),'footing_cover':float(values['footing_cover']),'H':float(values['H']),'D':float(values['D']),'tw':float(values['tw']),'Ltoe':float(values['Ltoe']),'Lheel':float(values['Lheel']),'h':h,'wall_cover':float(values['wall_cover'])}
            Results = Retaining_wall_script.bearing(Dimensions,float(values['Allowable_bearing']),float(values['DC']),float(values['Soil_density']),float(values['Surcharge']),float(values['Additional_moment']),float(values['Ka']),float(values['Internal_friction']),float(values['c']),float(values['Kp']),values,values['Top_restraint'],values['Restraint_height'])
            if values['Masonry'] == False:
                Results2 = Retaining_wall_script.concrete_wall(Dimensions,float(values['Allowable_bearing']),float(values['DC']),float(values['Soil_density']),float(values['Surcharge']),float(values['Additional_moment']),float(values['Ka']),float(values['Internal_friction']),float(values['c']),float(values['fc']),values['crack'],values)
                window['wall_vertical_steel'].update(str(round(Results2[0])))
                window['wall_horizontal_steel'].update(str(round(Results2[1])))
            else:
                Masonry = {'Masonry_size':values['Masonry_size'],'BT':float(values['BT']),'Masonry_type':values['Masonry_type'],'d':values['d']}
                Results2 = Retaining_wall_script.masonry_wall(float(values['d']),500,float(values['fuc']),Masonry,Dimensions,float(values['Ka']),float(values['Surcharge']),float(values['Soil_density']),values)

                if isinstance(Results2[0],str) == True:
                    window['wall_vertical_steel'].update(Results2[0])
                    window['wall_horizontal_steel'].update(Results2[1])
                else:
                    window['wall_vertical_steel'].update(str(round(Results2[0])))
                    window['wall_horizontal_steel'].update(str(round(Results2[1])))
            Results3 = Retaining_wall_script.concrete_heel(Dimensions, float(values['Allowable_bearing']),
                                                           float(values['DC']), float(values['Soil_density']),
                                                           float(values['Surcharge']),
                                                           float(values['Additional_moment']), float(values['Ka']),
                                                           float(values['Internal_friction']), float(values['c']),
                                                           float(values['fc']), values['crack'],values)
            Results4 = Retaining_wall_script.concrete_toe(Dimensions, float(values['Soil_density']),
                                                           float(values['Surcharge']),
                                                           float(values['Additional_moment']), float(values['Ka']),
                                                           float(values['Internal_friction']), float(values['c']),
                                                           float(values['fc']), values['crack'],Results['qmax_ULS'],Results['a_ULS'],values)
            if isinstance(Results['qmax'],str) == True:
                window['bearing_check'].update(Results['qmax'])
            else:
                window['bearing_check'].update(str(round(Results['qmax'])))
                if Results['qmax'] <= float(values['Allowable_bearing']):
                    window['bearing_OK/NG'].update('OK', text_color='green')
                else:
                    window['bearing_OK/NG'].update('NG', text_color='red')
            window['sliding_check'].update(str(round(Results['F'],1)))

            window['heel_horizontal_steel'].update(str(round(Results3[0])))
            window['heel_OoP'].update(str(round(Results3[1])))
            window['toe_horizontal_steel'].update(str(round(Results4[0])))
            window['toe_OoP'].update(str(round(Results4[1])))
            Results1 = Retaining_wall_script.overturning(Dimensions,float(values['Allowable_bearing']),float(values['DC']),float(values['Soil_density']),float(values['Surcharge']),float(values['Additional_moment']),float(values['Ka']),values,values['Top_restraint'],values['Restraint_height'])

            window['overturning_check'].update(str(round(Results1)))
            if Results1 <=0:
                window['overturning_OK/NG'].update('NG',text_color='red')
            else:
                window['overturning_OK/NG'].update('OK', text_color='green')
            if Results['F'] <=0:
                window['sliding_OK/NG'].update('NG',text_color='red')
            else:
                window['sliding_OK/NG'].update('OK', text_color='green')
        elif event == 'back':
            window.close()
            menu(variables)


def Wind():
    layout = [
        [sg.Text('Base Wind pressure calculator')],
        [sg.Text('Input')],
        [sg.Button('Back',key='back')]
    ]
    window = sg.Window('Wind Calcualtor',layout)

    while True:
        event, values = window.read()
        if event == sg.WIN_CLOSED:
            break
        elif event == 'back':
            window.close()
            menu(variables)


def Concrete_Beam_Layout(concrete_beam):
    try:
        x = concrete_beam['Shape']
    except:
        x = 'Rectangle'
    if x == 'Rectangle':
        layout = [
            [sg.Text('Concrete Calculator:')],
            [sg.Column([
                [sg.Text('Breadth, b:')],
                [sg.Text('Depth, D:')],
                [sg.Text('Cover:')],

            ]), sg.Column([
                [sg.Input(default_text= KeyCheck(concrete_beam,'breadth'), key='breadth', size=(5, 1),enable_events=True)],
                [sg.Input(default_text=KeyCheck(concrete_beam,'depth'), key='depth', size=(5, 1),enable_events=True)],
                [sg.Input(default_text=KeyCheck(concrete_beam,'cover'), key='cover', size=(5, 1),enable_events=True)]
            ]),
                sg.Column([
                    [sg.Text('mm')],
                    [sg.Text('mm')],
                    [sg.Text('mm')]
                ]),
                sg.Column([
                    [sg.Text('f\'c', )],
                    [sg.Text('')],
                    [sg.Text('')]
                ]),
                sg.Column([
                    [sg.Input(default_text=KeyCheck(concrete_beam,'fc'), key='fc', size=(5, 1),enable_events=True)],
                    [sg.Text('')],
                    [sg.Text('')]
                ]),
                sg.Column([
                    [sg.Text('MPa')],
                    [sg.Text('')],
                    [sg.Text('')]
                ]),
                sg.Column([
                    [sg.Text('Member Type:')],
                    [sg.Text('Member Shape')],
                    [sg.Text('')]
                ]),
                sg.Column([
                    [sg.Combo(['Two-way Slab', 'One-way Slab', 'Beam', 'Column'], size=(15, 1), enable_events=True,
                              default_value=KeyCheck(concrete_beam,'Member Type'), key='Member Type')],
                    [sg.Combo(['Rectangle','Circular'],default_value=KeyCheck(concrete_beam,'Shape'), key = 'Shape', size = (10,1), enable_events=True)],
                    [sg.Text('')]
                ])
            ],
            [sg.Column([
                [sg.Text('Rows of reinforcement:')],
            ]), sg.Column([
                [sg.Input(default_text=KeyCheck(concrete_beam,'number of rows of reinforcement'), key='number of rows of reinforcement', size=(3, 1),
                          enable_events=True)]
            ])]

        ]
    elif x == 'Circular':
        layout = [
            [sg.Text('Concrete Calculator:')],
            [sg.Column([
                [sg.Text('Diameter, b:')],

                [sg.Text('Cover:')],

            ]), sg.Column([
                [sg.Input(default_text=KeyCheck(concrete_beam, 'breadth'), key='breadth', size=(5, 1),
                          enable_events=True)],
                [sg.Input(default_text=KeyCheck(concrete_beam, 'cover'), key='cover', size=(5, 1), enable_events=True)]
            ]),
                sg.Column([
                    [sg.Text('mm')],
                    [sg.Text('mm')]
                ]),
                sg.Column([
                    [sg.Text('f\'c', )],
                    [sg.Text('')]
                ]),
                sg.Column([
                    [sg.Input(default_text=KeyCheck(concrete_beam, 'fc'), key='fc', size=(5, 1), enable_events=True)],
                    [sg.Text('')]
                ]),
                sg.Column([
                    [sg.Text('MPa')],
                    [sg.Text('')]
                ]),
                sg.Column([
                    [sg.Text('Member Type:')],
                    [sg.Text('Member Shape')],
                    [sg.Text('')]
                ]),
                sg.Column([
                    [sg.Combo(['Two-way Slab', 'One-way Slab', 'Beam', 'Column'], size=(15, 1), enable_events=True,
                              default_value=KeyCheck(concrete_beam, 'Member Type'), key='Member Type')],
                    [sg.Combo(['Rectangle', 'Circular'], default_value=KeyCheck(concrete_beam, 'Shape'), key='Shape',
                              size=(10, 1), enable_events=True)],
                    [sg.Text('')]
                ])
            ],
            [sg.Column([
                [sg.Text('Number of bars:')],
            ]), sg.Column([
                [sg.Input(default_text=KeyCheck(concrete_beam, 'number of rows of reinforcement'),
                          key='number of rows of reinforcement', size=(3, 1),
                          enable_events=True)]
            ])]

        ]
    D = 1000
    cover = 10
    for x in range(int(KeyCheck(concrete_beam,'number of rows of reinforcement'))):
        layout += [
            [sg.Column([[sg.Text('Row ' + str(x + 1) + ': Area of Steel:')]]),
             sg.Column([[sg.Input(default_text=KeyCheck(concrete_beam,'Ast' + str(x + 1)), key='Ast' + str(x + 1), size=(5, 1),enable_events=True)]]),
             sg.Column([[sg.Text('Depth to row centroid:')]]),
             sg.Column([[sg.Input(default_text=KeyCheck(concrete_beam,'y' + str(x + 1)), size=(5, 1), key='y' + str(x + 1),enable_events=True)]]),
             sg.Column([[sg.Text('Steel Strength:')]]),
             sg.Column([[sg.Input(default_text=KeyCheck(concrete_beam,'fy' + str(x + 1)), size=(5, 1), key='fy' + str(x + 1),enable_events=True)]]),
             sg.Column([[sg.Checkbox('By Area of steel')]])]
        ]
    layout += [
        [sg.Text('Area of steel:'), sg.Input(key='stirrups', default_text='10', size=(5, 1)), sg.Text('mm^2 Spacing:'),
         sg.Input(key='stirrup2', size=(5, 1), default_text='100'), sg.Text('fy:'),
         sg.Input(key='stirrup3', size=(5, 1), default_text='300')],
        [sg.Text('Demands:'),
         sg.Checkbox('Include in printout', key='print demands', default=True, enable_events=True)],
        [sg.Text('Moment:'), sg.Input(default_text=KeyCheck(concrete_beam,'Moment'), key='Moment', size=(5, 1),enable_events=True), sg.Text('KNm     Shear:'),
         sg.Input(default_text=KeyCheck(concrete_beam,'Shear'), key='Shear', size=(5, 1),enable_events=True), sg.Text('KN      Axial:'),
         sg.Input(default_text=KeyCheck(concrete_beam,'Axial'), key='Axial', size=(5, 1),enable_events=True), sg.Text('KN')],
        [sg.Text('Demands Calculator:'),
         sg.Checkbox('Include in printout', default=False, key='print demand calculator', enable_events=True)],
        [sg.Text('Results:')]]

    Results = [
            [sg.Column([
                [sg.Text('\u03A6Mu =')]
            ]),
                sg.Column([
                    [sg.Text('   KNm', key='M')]
                ]),
                sg.Column([
                    [sg.Text('', key='M_ch')]
                ]),
                sg.Column([
                    [sg.Text('\u03A6Mu,min =')]
                ]),
                sg.Column([
                    [sg.Text('   KN', key='Mu,min')]
                ]),
                sg.Column([
                    [sg.Text('', key='M_check')]
                ]),
                sg.Column([
                    [sg.Text('\u03A6Vuc =')]
                ]),
                sg.Column([
                    [sg.Text('   KN', key='Vuc')]
                ]),
                sg.Column([
                    [sg.Text('\u03A6Vus =')]
                ]),
                sg.Column([
                    [sg.Text('   KN', key='Vus')]
                ]),
                sg.Column([
                    [sg.Text('\u03A6Vu =')]
                ]),
                sg.Column([
                    [sg.Text('   KN', key='Vu')]
                ]),
                sg.Column([
                    [sg.Text('', key='Vu_check')]
                ])
            ],
            [sg.Text('Servicibility:'), sg.Text('Msls: '), sg.Input(default_text=KeyCheck(concrete_beam,'Msls'), key='Msls', size=(5, 1),enable_events=True),sg.Text('KNm')],
            [sg.Column([
                [sg.Text('Icr =')],
                [sg.Text('Ec =')]
            ]),
                sg.Column([
                    [sg.Text('   mm4', key='Icr')],
                    [sg.Text('   GPa', key = 'Ec')]
                ]),
                sg.Column([
                    [sg.Text('Ig = ')],
                    [sg.Text('Mcr.t = ')]
                ]),
                sg.Column([
                    [sg.Text('   mm4', key='Ig')],
                    [sg.Text('   KNm',key='Mcr.t')]
                ]),
                sg.Column([
                    [sg.Text('\u03C3scr = ')],
                    [sg.Text('Neutral axis = ')]
                ]),
                sg.Column([
                    [sg.Text('   MPa', key='steel_stress')],
                    [sg.Text('',key = 'dn')]
                ]),
                sg.Column([
                    [sg.Text('Failure Mode:')],
                    [sg.Text('Ief:')]
                ]),
                sg.Column([
                    [sg.Text('   ', key='Failure_mode')],
                    [sg.Text('   ', key='Ief')]
                ])
            ],
            [sg.Button('Calculate', key='Calculate')],
            [sg.Button('Print calculations', key='print_calcs')],
            [sg.Text('Type Job Name:'), sg.Input(default_text=KeyCheck(concrete_beam,'job_name'), key='job_name')],
            [sg.Text('Choose destination:'),
             sg.Input(key='print_location', default_text=KeyCheck(concrete_beam,'print_location')),
             sg.FolderBrowse()],
            [sg.Button('Back', key='back')],

        ]

    layout1 = layout + Results
    return layout1
def Concrete_Beam(concrete_beam):
    layout1 = Concrete_Beam_Layout(concrete_beam)
    window = sg.Window('Concrete Beam', layout1)
    while True:
        event, values = window.read()
        for i in values:
            if event == i:
                variable_write(values,'concrete_beam')
        if event == sg.WIN_CLOSED:
            break
        elif event == 'Member Type':

            if values['Member Type'] == 'Two-way Slab':

                concrete_beam_variable_write(values)
                concrete_beam = concrete_beam_variable()
                layout1 = Concrete_Beam_Layout(concrete_beam)
                window2 = sg.Window('Steel calculator', layout1)
                window.close()
                window = window2
                #Two_way = Concrete.slab(values['G'],values['Q'])
            elif values['Member Type'] == 'One-way Slab' or values['Member Type'] == 'Beam':
                layout1 = Concrete_Beam_Layout(concrete_beam)
                window2 = sg.Window('Steel calculator', layout1)
                window.close()
                window = window2
        elif event == 'Shape' and values['Shape'] == 'Circular' or event == 'Shape' and values['Shape'] == 'Rectangle':
            concrete_beam = variable('concrete_beam')
            layout1 = Concrete_Beam_Layout(concrete_beam)
            window2 = sg.Window('Steel calculator', layout1)
            window.close()
            window = window2
        elif event == 'print_calcs':
            y = [0 for x in range(concrete_beam['number of rows of reinforcement'])]
            Ast = [0 for x in range(concrete_beam['number of rows of reinforcement'])]
            fy = [0 for x in range(concrete_beam['number of rows of reinforcement'])]
            values['stirrups'] = [float(values['stirrups']), float(values['stirrup2']), float(values['stirrup3'])]
            for x in range(concrete_beam['number of rows of reinforcement']):
                y[x] = float(values['y' + str(x + 1)])
                Ast[x] = float(values['Ast' + str(x + 1)])
                fy[x] = float(values['fy' + str(x + 1)])
            if values['Member Type'] == 'Two-way Slab':
                Two_way = Concrete.slab(float(values['G']),float(values['Q']),float(values['L1'])*1000,float(values['L2'])*1000,[1,2,3,4,5,6,7,8,9],Ast[0],float(values['fc']),int(values['depth']),float(values['cover']))
            elif values['Member Type'] == 'One-way Slab' or values['Member Type'] == 'Beam':
                One_way = Concrete.One_way_slab_print(int(values['breadth']),int(values['depth']),int(values['cover']),y,Ast,float(values['Moment']),float(values['Shear']),float(values['Axial']),values['stirrups'],float(values['fc']),fy,
                                                      float(values['G']),float(values['Q']),float(values['L']),values['print_location'],values['job_name'])
        elif event == 'Calculate':
            concrete_beam['number of rows of reinforcement'] = int(concrete_beam['number of rows of reinforcement'])
            y = [0 for x in range(concrete_beam['number of rows of reinforcement'])]
            Ast = [0 for x in range(concrete_beam['number of rows of reinforcement'])]
            fy = [0 for x in range(concrete_beam['number of rows of reinforcement'])]
            values['stirrups'] = [float(values['stirrups']),float(values['stirrup2']),float(values['stirrup3'])]
            for x in range(concrete_beam['number of rows of reinforcement']):
                y[x] = float(values['y'+str(x+1)])
                Ast[x] = float(values['Ast' + str(x+1)])
                fy[x] = float(values['fy'+str(x+1)])
            print(y, 'main x')
            try:
               depth =  int(values['depth'])
            except:
                depth = int(values['breadth'])
            result = Concrete.beam_moment(int(values['breadth']),depth,int(values['cover']),y,Ast,float(values['Moment']),float(values['Shear']),float(values['Axial']),values['stirrups'],float(values['fc']),fy,values['Shape'])
            for x in range(concrete_beam['number of rows of reinforcement']):
                y[x] = float(values['y'+str(x+1)])
                Ast[x] = float(values['Ast' + str(x+1)])
            print(y, 'main y')
            result1 = Concrete.beam_shear(int(values['breadth']), depth, int(values['cover']), y, Ast,
                                         float(values['Moment']), float(values['Shear']), float(values['Axial']),
                                          values['stirrups'], float(values['fc']), fy,values['Shape'])
            result2 = Concrete.service_moments(int(values['breadth']), depth, int(values['cover']), y, Ast,
                                         float(values['Moment']), float(values['Shear']), float(values['Axial']),
                                          values['stirrups'], float(values['fc']), fy,values)
            window['M'].update(str(round(result['Mn']*0.65,2)) + '   KNm')
            window['Vuc'].update(str(round(result1['Vuc']/1000*0.6,2)) + '   KN')
            window['Vus'].update(str(round(result1['Vus'] / 1000*0.6, 2)) + '   KN')
            window['Vu'].update(str(round((result1['Vus'] + result1['Vuc'])/1000*0.6, 2)) + '   KN')
            window['Icr'].update(str("{:.3e}".format(result2['Icr'])) + ' mm^4')
            window['Mcr.t'].update(str(round(result2['Mcr.t'],1)) + ' KNm')
            window['steel_stress'].update(str(round(result2['steel_stress'])) + ' MPa')
            window['Mu,min'].update(str(round(result['Mu_min']*0.65,1))+ ' KNm')
            window['Ec'].update(str(round(result2['Ec'])) + ' GPa')
            window['Ig'].update(str("{:.3e}".format(result2['Ig'])) + ' mm4')
            window['dn'].update(str(round(result2['dn'])) + ' mm')
            window['Ief'].update(str("{:.3e}".format(result2['Ief'])) + ' mm4')
            if result['Failure mode'] == 'Tensile failure':
                window['Failure_mode'].update(result['Failure mode'],text_color='green')
            else:
                window['Failure_mode'].update(result['Failure mode'], text_color='red')
            if float(values['Moment']) < result['Mn']*0.65 and result['Mn'] > result['Mu_min']:
                window['M_check'].update('OK',text_color='green')
            else:
                window['M_check'].update('NG',text_color='red')
            if float(values['Shear']) > (result1['Vus'] + result1['Vuc'])*0.6:
                window['Vu_check'].update('NG',text_color='red')
            else:
                window['Vu_check'].update('OK',text_color='green')
        elif event == 'number of rows of reinforcement':
            for i in range(int(concrete_beam['number of rows of reinforcement']),int(values['number of rows of reinforcement'])):
                values['y' + str(i+1)] = 100
                values['Ast' + str(i + 1)] = 0
                values['fy' + str(i + 1)] = 500
            variable_write(values,'concrete_beam')
            concrete_beam = variable('concrete_beam')
            window.close()
            #y = [values['y' + str(x + 1)] for x in range(concrete_beam['rows'])]
            #Ast = [values['Ast' + str(x + 1)] for x in range(concrete_beam['rows'])]
            Concrete_Beam(concrete_beam)
        elif event == 'back':
            window.close()
            variables['concrete']['b'] = values['breadth']
            variables['concrete']['d'] =values['depth']
            variables['concrete']['cover'] = values['cover']
            variables['rows'] = int(values['number of rows of reinforcement'])
            variables['concrete']['fc'] = values['fc']
            y = [values['y'+str(x+1)] for x in range(concrete_beam)]
            Ast = [values['Ast' + str(x + 1)] for x in range(concrete_beam)]
            variables['concrete']['y'] = y
            variables['concrete']['Ast'] = Ast
            menu(variables)
#breadth,depth,cover,rows,fc,y,Ast,stirrups


def Timber():
    layout = [[sg.Column([[sg.Text('Timber Calculator')],
              [sg.Input(key = 'b',size=(5,1),default_text='45'),sg.Text('Breadth')],
              [sg.Input(key = 'd',size = (5,1),default_text='240'),sg.Text('Depth')],
              [sg.Combo(['MGP 10', 'MGP 12','MGP 15', 'meySPAN 13','F17','F27'],key = 'grade',default_value='MGP 10',size=(15,1)),sg.Text('Grade')],
              [sg.Combo(['Seasoned','Unseasoned'],default_value='Seasoned',key = 'Seasoned'),sg.Text('Seasoned')],
              [sg.Combo(['5 seconds', '5 minutes', '5 hours', '5 days','5 months','50+ years'],default_value='5 months',key = 'load_duration'),sg.Text('Load Duration')],
              [sg.Combo(['Yes','No'],key='nailed',size=(5,1),default_value='Yes'),sg.Text('Are members nailed together?')],
              [sg.Combo([1,2,3,4,5,6,7,8,9],default_value=1,key = 'ncom'),sg.Text('Number of Members joined')],
              [sg.Combo([1,2,3,4,5,6,7,8,9],key = 'nmem',default_value=1),sg.Text('Number of parallel member groups separated by spacing s')],
              [sg.Input(key='s',size=(5,1),default_text='1'),sg.Text('Spacing, s in mm, if not applicable ignore')],
              [sg.Input(key='Length',size=(5,1),default_text='1000'),sg.Text('Length in mm'),sg.Input(key = 'segment length',size=(5,1)),sg.Text('Segment Length')],
              [sg.Button('Calculate',key = 'Calculate')],
              [sg.Text(key='Md')],
              [sg.Text(key='Vd')],
              [sg.Text('Filename:')],
              [sg.Input(default_text='TimberBeam',key = 'name')],
              [sg.Text('Comments:')],
              [sg.Input(key='comments',default_text='')],
              [sg.Text('Choose destination:'), sg.Input(key='print_location',default_text=r'C:\Users\tduffett\PycharmProjects\pythonProject1'), sg.FolderBrowse()],
              [sg.Button('Print Calcs',key='print'),sg.Checkbox('Include Design Actions',default=True, key='Design Actions')],
              [sg.Button('Back',key='back')]]),sg.VSeparator(),sg.Column([[sg.Text('Simply Supported:\nDisplacement Calculator\nDemand Calculator')],
                                                                          [sg.Input(key = 'G_UDL',size=(5,1),default_text='0'),sg.Text('G UDL'),sg.Input(key = 'Q_UDL',size = (5,1),default_text='0'),sg.Text('Q UDL')],
                                                                          [sg.Input(key = 'G_PL',size=(5,1),default_text='0'),sg.Text('G PL'),sg.Input(key = 'Q_PL',size = (5,1),default_text='0'),sg.Text('Q PL')],
                                                                          [sg.Input(key= 'distance',size=(5,1),default_text='0'),sg.Text('Point load distance from support')],
                                                                          [sg.Button('Calculate',key = 'deflection')],
                                                                          [sg.Column([
                                                                              [sg.Text(key='deflection results')]
                                                                          ]),sg.VSeparator(),sg.Text(key='moment results'),sg.VSeparator(),sg.Text(key='shear results')]]
                                                                         )]]
    window = sg.Window('Timber',layout)

    while True:
        event, values = window.read()
        if event == 'Calculate':
            if values['Seasoned'] == 'Seasoned':
                seasoned = True
            else:
                seasoned = False
            if values['nailed'] == 'Yes':
                nailed = True
            elif values['nailed'] == 'No':
                nailed = False
            Md = beam_moment(float(values['b']),float(values['d']),values['load_duration'],seasoned,int(values['ncom']),int(values['nmem']),int(values['s']),int(values['Length']),values['grade'],nailed)
            Vd = beam_shear(float(values['b']), float(values['d']), values['load_duration'], seasoned, int(values['ncom']),
                        int(values['nmem']), int(values['s']), int(values['Length']), values['grade'])
            window['Md'].update('PhiMd = '+ str(round(Md[0]/1000,1)) + ' KNm')
            window['Vd'].update('PhiVd = ' + str(round(Vd[0]/1000,1)) + ' KN')
        elif event == sg.WIN_CLOSED:
            break
        elif event == 'print':
            if values['Seasoned'] == 'Seasoned':
                seasoned = True
            else:
                seasoned = False
            if values['nailed'] == 'Yes':
                nailed = True
            elif values['nailed'] == 'No':
                nailed = False
            Md = beam_moment(float(values['b']), float(values['d']), values['load_duration'], seasoned,
                             int(values['ncom']), int(values['nmem']), int(values['s']), int(values['Length']),
                             values['grade'], nailed)
            Vd = beam_shear(float(values['b']), float(values['d']), values['load_duration'], seasoned,
                            int(values['ncom']),
                            int(values['nmem']), int(values['s']), int(values['Length']), values['grade'])
            d = deflection_check(float(values['b']), float(values['d']), int(values['ncom']), values['grade'],
                                 float(values['G_UDL']), float(values['Q_UDL']), float(values['G_PL']),
                                 float(values['Q_PL']), float(values['Length']))
            print_calcs(values['name'],values['Design Actions'],d[2]+ d[4],d[3]+d[5],Md[0],Md[1],Md[2],Md[3],Md[4],Md[5],Md[6],Md[7],Vd[1],Vd[2],Md[8],Vd[0],values['print_location'],values['comments'],d[0],d[1],float(values['Length']))
        elif event == 'deflection':
            d = deflection_check(float(values['b']),float(values['d']),int(values['ncom']),values['grade'],float(values['G_UDL']),float(values['Q_UDL']),float(values['G_PL']),float(values['Q_PL']),float(values['Length']))
            window['deflection results'].update('UDL Deflection: '+ str(round(d[0],2)) + ' mm\nPoint Load Deflection: ' + str(round(d[1],2)) +
                                                ' mm\nTotal Deflection: ' + str(round(d[0] + d[1],2)) + ' mm')
            window['moment results'].update('UDL Moment:' + str(round(d[2],2))+ ' KNm\nPL Moment: ' + str(round(d[4],2)) + ' KNm\nTotal Moment :' +
                                                str(round(d[2] + d[4],2)) + ' KNm')
            window['shear results'].update('UDL Shear: '+
                                                str(round(d[3],2)) + ' KN\nPL Shear :' + str(round(d[5],2)) + ' KN\nTotal Shear :' + str(round(d[3] + d[5],2)) + ' KN')
        elif event == 'back':
            window.close()
            menu(variables)
        elif event == sg.WIN_CLOSED:
            break

def steel_calculator():

    # load in data to be read
    from openpyxl import load_workbook

    wb = load_workbook(filename='Steel Design Calculator.xlsx')
    # SectionType = wb.sheetnames
    SectionSize = []
    # print(sheet_ranges['A6'].value)
    SectionType = ['Universal_Beam', 'Universal_Column', 'PFC', 'RHS','SHS','CHS', 'T-section']
    for j in wb['Universal_Beam']:
        SectionSize.append(j[0].value)
    SectionSize = list(filter(lambda item: item is not None, SectionSize))
    print(SectionSize)
    for j1 in range(4):
        SectionSize.pop(0)
    print(SectionSize)
    # load in GUI library


    # ttk_style = 'vista'
    layout = [
        [sg.Column([
            [sg.Text("Steel calculator", key='title')],
            [sg.Text('This calculator is not finished yet', key='t1')],
            [sg.Text('Available section types are Universal Columns, Universal Beams, and PFC\'s and RHS')],
            [sg.Text('Choose Section Type:')],
            [sg.Combo(SectionType, key='SectionType', enable_events=True, default_value=SectionType[0], size=(30, 1))],
            [sg.Text('Choose Section Size:')],
            [sg.Combo(SectionSize, key='SectionSize', default_value=SectionSize[0], size=(30, 1))],
            [sg.Text('Input Total Length in metres below:', key='t2')],
            [sg.Input(key='Length', size=(5, 1), default_text='1')],
            [sg.Text('Input segment Length in metres below:')],
            [sg.Input(default_text='1', key='segment length', size=(5, 1))],
            [sg.Text('Input alpha m below:', key='t3')],
            [sg.Input(key='alpha_m', size=(5, 1), default_text='1')],
            [sg.Combo(['FF', 'FP', 'FL', 'PP', 'PL', 'LL', 'FU', 'PU'], key='restraint', default_value='FP',
                      enable_events=True)],
            [sg.Combo(['Shear centre', 'Top flange'], key='load_height_position', default_value='Shear centre',
                      size=(15, 1))],
            [sg.Combo(['Within segment', 'At segment end'], key='longitudinal_position', default_value='Within segment',
                      size=(15, 1)),sg.Text('Load height')],
            [sg.Combo(['Any', 'None', 'One', 'Both'], key='ends_with_restraint', default_value='One'),sg.Text('Ends with Lateral restraint')],
            [sg.Button('Calculate', key='calculate', use_ttk_buttons=True)],
            [sg.Button('Back', key='back'), sg.Button('quit', key='b2', use_ttk_buttons=True)],
            [sg.Column([
                [sg.Text('\u03A6Msx = '),sg.Text('', key='PhiMsx')],
                [sg.Text('\u03A6Mbx = '),sg.Text('', key='PhiMbx')],
                [sg.Text('\u03A6Msy = '),sg.Text('', key='PhiMsy')],
                [sg.Text('\u03A6Mby = '),sg.Text('', key='PhiMby')]
            ]),sg.Column([
                [sg.Text('\u03A6Nsx = '),sg.Text('', key = 'PhiNsx')],
                [sg.Text('\u03A6Ncx = '),sg.Text('',key = 'PhiNcx')],
                [sg.Text('\u03A6Nsy = '),sg.Text(key = 'PhiNsy')],
                [sg.Text('\u03A6Ncy = '),sg.Text(key = 'PhiNcy')]

            ]),sg.Column([
                [sg.Text('\u03A6Vu = '), sg.Text('', key='PhiVu')],
                [sg.Text('\u03A6Vvm = '), sg.Text('', key='PhiVvm')]
            ])],
            [sg.Button('Print calculations', key='print_calcs')],
            [sg.Text('Type Job Name:'), sg.Input(default_text='610UB125', key='job_name')],
            [sg.Text('Choose destination:'),
             sg.Input(key='print_location', default_text=r'C:\Users\tduffett\PycharmProjects\pythonProject1'),
             sg.FolderBrowse()]]),
            sg.VSeparator()

           # ,sg.Column(
            #[
             #   [sg.Canvas(key='canvas')],
             #   [sg.Button('Calculate',key='deflection'),sg.Text('Point Loads:'),sg.Input(default_text='1 1',key = 'PL',size=(10,1))
             #       ,sg.Text('format:[Load,position,etc.]   Length'),sg.Input(default_text='2 2',key = 'L',size=(10,1)),
             #    sg.Text('format: [span1,span2,etc.]   UDL:'),sg.Input(default_text='1 2 3',key = 'UDL',size=(10,1)),sg.Text('format:[start,end,load]')]
            #]
            #)
        ]]


    # Create the window
    window = sg.Window("Steel Calculator", layout, resizable=True)

    # Create an event loop
    while True:
        event, values = window.read()
        if event == 'calculate':
            SectionType1 = values['SectionType']
            SectionSize1 = values['SectionSize']

            alpha_m = float(values['alpha_m'])
            restraint = values['restraint']
            load_height_position = values['load_height_position']
            longitudinal_position = values['longitudinal_position']
            ends_with_restraint = values['ends_with_restraint']
            section_properties = st.section_properties(SectionType1, SectionSize1,0,0,0)
            st.compact(section_properties, SectionType1, float(values['segment length']), alpha_m, restraint, load_height_position,
                                      longitudinal_position, ends_with_restraint)
            st.shear(section_properties,SectionType1)
            st.shear_moment(section_properties,0)
            st.axial_compression(section_properties,SectionType1,float(values['segment length']))

            print(section_properties)
            # End program if user closes window or
            # presses the OK button
            # if event == 'quit' or event == sg.WIN_CLOSED:
            window['PhiMsx'].update(str(round(section_properties['PhiMsx'], 1)) + 'kNm')

            window['PhiMbx'].update(str(round(section_properties['PhiMbx'], 1)) + 'kNm')
            window['PhiMsy'].update(str(round(section_properties['PhiMsy'],1)) + 'kNm')
            window['PhiMby'].update(str(round(section_properties['PhiMby'], 1)) + 'kNm')
            window['PhiNsx'].update(str(round(section_properties['PhiNsx'], 1)) + 'kN')
            window['PhiNcx'].update(str(round(section_properties['PhiNcx'], 1)) + 'kN')
            window['PhiNsy'].update(str(round(section_properties['PhiNsy'], 1)) + 'kN')
            window['PhiNcy'].update(str(round(section_properties['PhiNcy'], 1)) + 'kN')
            window['PhiVu'].update(str(round(section_properties['PhiVu'], 2)) + ' KN')
        elif event == 'calc_T_section':
            section_properties = st.section_properties(values['SectionType'],0,values['b'],values['d'],values['t'])
            calculations = st.compact(section_properties,values['SectionType'], float(values['Length']), float(values['alpha_m']),values['restraint'],
                                      values['load_height_position'], values['longitudinal_position'],values['ends_with_restraint'])
            window['result'].update('PhiMsx = ' + str(round(calculations['PhiMsx'], 1)) + 'kNm  Actual capacity ~90% of shown here')
            window['PhiMsx'].update('PhiMbx = ' + str(round(calculations['PhiMbx'], 1)) + 'kNm  Actual capacity ~90% of shown here')
        elif event == 'dLimit':
            for x in values['dLimit'].split('/'):
                if x.isdigit():
                    Limit = int(x)
            window['dLimit1'].update(str(round(float(values['Length'])*1000/Limit,2)) + ' mm')
            if float(values['Length'])*1000/Limit > d[0]+d[1]:
                window['dCheck'].update('OK')
            else:
                window['dCheck'].update('NG')
        elif event == 'deflection':
            PL = list(map(float,values['PL'].strip().split()))
            L = list(map(float,values['L'].strip().split()))
            UDL = list(map(float, values['UDL'].strip().split()))
            fig_1 = Deflection.any_spans(UDL,PL,L,1)
            fig = matplotlib.figure.Figure(figsize=(10, 4), dpi=100)
            fig.add_subplot(111).plot(fig_1[0],fig_1[2],fig_1[0],[0]*3999,fig_1[0],fig_1[1])

            fig_canvas_agg = draw_figure(window['canvas'].TKCanvas, fig)

        elif event == 'SectionType':
            if values[event] == 'T-section':
                layout = [
                    [sg.Text("Steel calculator", key='title')],
                    [sg.Text('This calculator is not finished yet', key='t1')],
                    [sg.Text('Available section types are Universal Columns, Universal Beams, and PFC\'s and RHS')],
                    [sg.Text('Choose Section Type:')],
                    [sg.Combo(SectionType, key='SectionType', enable_events=True, default_value='T-section',
                              size=(30, 1)), sg.Text('fy'), sg.Input(key='fyf', default_text='300', size=(5, 1))],
                    [sg.Text('Choose Section Size:')],
                    [sg.Text('b'), sg.Input(key='b', enable_events=True, default_text='89',
                                            size=(5, 1)), sg.Text('d'),
                     sg.Input(key='d', enable_events=True, default_text='89',
                              size=(5, 1)), sg.Text('t'),
                     sg.Input(key='t', enable_events=True, default_text='5',
                              size=(5, 1)),sg.Text('w'),sg.Input(key = 'w',default_text='10', size = (5,1))],
                    [sg.Text('Input segment Length in metres below:', key='t2')],
                    [sg.Column([[sg.Input(key='Length', size=(5, 1), default_text='1')],
                    [sg.Text('Input alpha m below:', key='t3')],
                    [sg.Input(key='alpha_m', size=(5, 1), default_text='1')],
                    [sg.Combo(['FF', 'FP', 'FL', 'PP', 'PL', 'LL', 'FU', 'PU'], key='restraint', default_value='FP',
                              enable_events=True)],
                    [sg.Combo(['Shear centre', 'Top flange'], key='load_height_position', default_value='Shear centre',
                              size=(15, 1))],
                    [sg.Combo(['Within segment', 'At segment end'], key='longitudinal_position',
                              default_value='Within segment',
                              size=(15, 1))],
                    [sg.Combo(['Any', 'None', 'One', 'Both'], key='ends_with_restraint', default_value='One')],
                    [sg.Button('Calculate', key='calc_T_section', use_ttk_buttons=True)],
                    [sg.Button('Back', key='back'), sg.Button('quit', key='b2', use_ttk_buttons=True)],
                    [sg.Text('', key='result'), sg.Text('',key = 'shear')],
                    [sg.Text('', key='PhiMsx')]]),sg.VSeparator(),sg.Image('Tsection.PNG')],
                    [sg.Button('Print calculations', key = 'print_calcs')],
                    [sg.Text('Type Job Name:'), sg.Input(default_text=values['SectionSize'], key = 'job_name')],
                    [sg.Text('Choose destination:'), sg.Input(key='print_location',
                                                              default_text=r'C:\Users\tduffett\PycharmProjects\pythonProject1'), sg.FolderBrowse()]
                ]
                window2 = sg.Window('Steel calculator', layout)
                window.close()
                window = window2
            elif values[event] == 'Universal_Beam' or values[event] == 'Universal_Column' or values[event] == 'PFC' or values[event] == 'RHS' or values[event]=='SHS' or values[event]=='CHS':
                item = values[event]
                matrix = []
                SectionNames = wb[item]
                for x in SectionNames:
                    matrix.append(x[0].value)
                for i in range(7):
                    matrix.pop(0)
                SectionSize = list(filter(lambda item: item is not None, SectionSize))
                window['SectionSize'].update(value=matrix[7], values=matrix)

        # window['b1'].update('OK')
        elif event == 'b2' or event == sg.WIN_CLOSED:
            break
        elif event == 'restraint':
            if values['restraint'] == 'FU' or values['restraint'] == 'PU':
                window['ends_with_restraint'].update(values = ['Any'], value = 'Any')
            elif values['restraint'] == 'FF' or values['restraint'] == 'FP' or values['restraint'] == 'PP':
                window['ends_with_restraint'].update(values = ['None', 'One', 'Both'], value = 'None')
            elif values['restraint'] == 'FL' or values['restraint'] == 'PL' or values['restraint'] == 'LL':
                window['ends_with_restraint'].update(values = ['None'], value = 'None')

        elif event == 'print_calcs':
            SectionType1 = values['SectionType']
            SectionSize1 = values['SectionSize']
            Length = float(values['Length'])
            alpha_m = float(values['alpha_m'])
            restraint = values['restraint']
            load_height_position = values['load_height_position']
            longitudinal_position = values['longitudinal_position']
            ends_with_restraint = values['ends_with_restraint']
            section_properties = st.section_properties(SectionType1, SectionSize1, 0, 0, 0)
            calculations = st.compact(section_properties, SectionType1, Length, alpha_m, restraint,
                                      load_height_position,
                                      longitudinal_position, ends_with_restraint)
            st.printcalcs(SectionSize1,section_properties,values['print_location'],values['job_name'],'',0,0,0,True)
        elif event == 'back':
            window.close()
            menu(variables)
    window.close()





def bolts():
    layout = [[sg.Text('Number of rows')],
              [sg.Input(key='#rows', enable_events=True, size=(5, 1)), sg.Button('OK', key='rows')],
              [sg.Button('Back', key = 'back')]
              ]
    window = sg.Window("Bolt Calculator", layout)

    while True:
        event, values = window.read()
        if event == 'rows':

            left_column = [[sg.Text('x')]] + [[sg.Input(size=(5, 1), key='x' + str(i), default_text='0')] for i in
                                              range(int(values['#rows']))]

            right_column = [[sg.Text('y')]] + [[sg.Input(size=(5, 1), key='y' + str(i)),
                                                sg.Input(key='Bolts in row' + str(i), size=(2, 1), default_text='1'),
                                                sg.Text('row' + str(i + 1))] for i in range(int(values['#rows']))]

            layout = [
                [sg.Text('Number of rows')],
                [sg.Input(default_text=values['#rows'], key='#rows', enable_events=True, size=(5, 1)),
                 sg.Button('OK', key='rows')],
                [sg.Text('Bolt Size'), sg.Input(default_text='20', key='bolt_size', size=(5, 1)), sg.Text('Bolt grade'),sg.Combo(['4.6/S'],default_value='4.6/S', key = 'grade')],
                [sg.Column(left_column, vertical_alignment='top'), sg.VSeparator(),
                 sg.Column(right_column, vertical_alignment='top')],
                [sg.Input(default_text='0', key='Moment', size = (5,1)), sg.Text('Moment')],
                [sg.Input(default_text='0', key='Shear', size = (5,1)), sg.Text('Shear')],
                [sg.Input(default_text='0', key='Axial', size = (5,1)), sg.Text('Axial (Tension Negative)')],
                [sg.Button('Calculate', key='Calculate'),sg.Button('Back', key = 'back')],
                [sg.Text('', key='result')],
                [sg.Text('', key='result2')],
                [sg.Text('', key = 'result3'),sg.Text('', key = 'result33')],
                [sg.Text('', key='result4'),sg.Text('', key = 'result44')]
            ]
            window2 = sg.Window('Bolt Calculator', layout)
            window.close()
            window = window2
        elif event == sg.WIN_CLOSED:
            break
        elif event == 'Calculate':
            bolts = {}
            bolts['Size'] = values['bolt_size']
            bolts['Moment'] = values['Moment']
            bolts['Shear'] = values['Shear']
            bolts['Axial'] = values['Axial']
            bolts['rows'] = values['#rows']
            bolts['grade'] = values['grade']
            for i in range(int(values['#rows'])):
                bolts['x' + str(i + 1)] = values['x' + str(i)]
                bolts['y' + str(i + 1)] = values['y' + str(i)]
                bolts['row_' + str(i + 1) + '_bolts'] = values['Bolts in row' + str(i)]
            result = bc.bolt_actions(bolts)
            window['result'].update(str(result[0]) + ' kN Axial Force on bolt')
            window['result2'].update(str(result[1]) + ' kN Shear Force on bolt')
            window['result3'].update(str(result[2]) + ' kN Bolt tensile capacity')
            window['result4'].update(str(result[3]) + ' kN Bolt shear capacity')
            if abs(result[0]) > result[2]:
                window['result33'].update('FAIL')

            elif abs(result[0]) < result[2]:
                window['result33'].update('PASS')

            if result[1] > result[3]:
                window['result44'].update('FAIL')

            elif result[1] < result[3]:
                window['result44'].update('PASS')

        elif event == 'back':
            window.close()
            menu(variables)
    window.close()

def development():
    right = [
        [sg.Text('fsy characteristic yield strength of reinforcement, determined in accordance with Clause 3.2.1')],
        [sg.Text('f \'c characteristic compressive (cylinder) strength of concrete at 28 days')],
        [sg.Text('Bar diameter in mm')],
        [sg.Text('k1 a coefficient that accounts for the bond properties of the bonded reinforcement (see Clause 8.6.2.3)')],
        [sg.Text('cd a dimension (in mm) corresponding to the smaller of the concrete'
                 ' cover to a bar developing stress and half the clear distance to the next'
                 ' parallel bar developing stress, as shown in Figure 13.1.2.2')]
    ]
    left = [
        [sg.Input(size=(5, 2), key='fsy', default_text='400')],
        [sg.Input(size=(5, 1), key='fc', default_text='20')],
        [sg.Input(size=(5, 1), key='db', default_text='30')],
        [sg.Input(size=(5, 1), key='k1', default_text=1)],
        [sg.Input(size=(5, 3), key='cd', default_text='30')]
    ]
    layout2 = [
        [sg.Text('Enter properties')],
        [sg.Column(left, vertical_alignment='top'), sg.VSeparator(), sg.Column(right, vertical_alignment='top')],
        [sg.Button('Calculate', key='calculate')],
        [sg.Text('Minimum Development Length (AS3600:2018 Clause 13.1.2.2):')],
        [sg.Text('', key='result')],
        [sg.Text('Bar stress developed over length:'), sg.Input(default_text='500',size=(5,1), key = 'L'),sg.Button('Calculate', key = 'calc2')],
        [sg.Text('', key = 'result1')],
        [sg.Button('Back', key='back')]
    ]
    window = sg.Window('Development Length Calculator', layout2, resizable=True)

    while True:
        event, values = window.read()
        if event == sg.WIN_CLOSED:
            break
        elif event == 'calculate':
            bar_properties = {}
            bar_properties['fsy'] = values['fsy']
            bar_properties['fc'] = values['fc']
            bar_properties['k1'] = values['k1']
            bar_properties['db'] = values['db']
            bar_properties['cd'] = values['cd']
            development_length = bc.Development_length(bar_properties)
            window['result'].update(str(development_length) + ' mm')
        elif event == 'calc2':
            bar_properties = {}
            bar_properties['fsy'] = values['fsy']
            bar_properties['fc'] = values['fc']
            bar_properties['k1'] = values['k1']
            bar_properties['db'] = values['db']
            bar_properties['cd'] = values['cd']
            development_length = bc.Development_length(bar_properties)
            window['result'].update(str(development_length) + ' mm')
            window['result1'].update(str(min(round(int(values['L'])/development_length*int(values['fsy'])),int(values['fsy']))) +
                                     ' MPa or ' +
                                     str(round(0.8*min(int(values['L'])/development_length*int(values['fsy']),int(values['fsy']))*int(values['db'])**2 / 4 * math.pi/1000)) + ' KN')
        elif event == 'back':
            window.close()
            menu(variables)
    window.close()

steel = 1
def draw_figure(canvas, figure):
    figure_canvas_agg = FigureCanvasTkAgg(figure, canvas)
    figure_canvas_agg.draw()
    figure_canvas_agg.get_tk_widget().pack(side='top', fill='both', expand=1)
    return figure_canvas_agg

concrete_beam = variable('concrete_beam')
variables = {'concrete':concrete_beam,'steel':steel}
menu(variables)