#This script contains functions to be called into the main script
from openpyxl import load_workbook
import math
from pylatex import Document, Section, Subsection, Command
from pylatex import  Tabular, Math, TikZ, Axis, \
    Plot, Figure, Matrix, Alignat, MultiRow, MultiColumn,Tabularx, TextColor
from pylatex import Document, PageStyle, Head, MiniPage, Foot, LargeText, \
    MediumText, LineBreak, simple_page_number, StandAloneGraphic, LongTabu, NewPage
from pylatex.utils import italic, NoEscape, bold
import subprocess
import os
wb = load_workbook(filename='Steel Design Calculator.xlsx')
#The first function will return data about the section selected.
def section_properties(SectionType, SectionSize, b,d,t):
    data = ['d', 'bf', 'tf', 'tw', 'Ix', 'Zx', 'Sx', 'rx', 'Iy', 'Iw', 'J','Zy','Sy','fyf','fyw','kf','Ag','ry']
    matrix = {}
    if SectionType == 'RHS' or SectionType == 'SHS':
        data = ['d', 'bf', 'tf', 'tw', 'Ix', 'Zx', 'Sx', 'rx', 'Iy', 'Iw', 'J','Zy','Sy','Zex','Zey','fyf','fyw','kf','Ag','ry']
    if SectionType == 'T-section':
        matrix['d'] = 1
    if SectionType == 'CHS':
        data = ['Ag','I','Z','S','r','J','Ze','fy','kf']

            #Use openpyxl package
            #Also needs to have relevant standard tables
            # First need to import section properties from database
    sheet_ranges = wb[SectionType]
            #use format x = sheet_ranges[A6].value

    count4 = 0
    count3 = 0
    for row1 in sheet_ranges:
        count3 +=1
        for cell1 in row1:
            if SectionSize == cell1.value:
                count4 = count3
                break
    count = 0
    for row in sheet_ranges:
        count += 1
        count1 =0
        for cell in row:
            count1 += 1
            for x in range(len(data)):
                if cell.value == data[x]:
                    data1 = sheet_ranges.cell(row=count4, column=count1)
                    matrix[cell.value] = (data1.value)
    if 'fyw' in matrix:
        matrix['fy'] = min(matrix['fyw'],matrix['fyf'])
    print(matrix)
    return matrix


def axial_compression(section_properties,SectionType,L):
    if SectionType != 'CHS':
        axial_compression_formula(section_properties,L,'x',SectionType)
        axial_compression_formula(section_properties,L,'y',SectionType)
    elif SectionType == 'CHS':
        axial_compression_formula(section_properties,L,'',SectionType)
        section_properties['PhiNsx'] = section_properties['PhiNs']
        section_properties['PhiNsy'] = section_properties['PhiNs']
        section_properties['PhiNcx'] = section_properties['PhiNc']
        section_properties['PhiNcy'] = section_properties['PhiNc']
    print('mark',section_properties)
    return section_properties


def axial_compression_formula(section_properties,L,xy,SectionType):
    section_properties['Ns'+ xy] = section_properties['kf']*section_properties['Ag']*section_properties['fy']/1000
    section_properties['PhiNs'+ xy] = 0.9*section_properties['Ns'+xy]
    ke = 1
    Le = L*ke*1000
    lambdan = (Le/section_properties['r' + xy])*math.sqrt(section_properties['kf'])*math.sqrt(section_properties['fy']/250)
    alpha_a = (2100*(lambdan-13.5))/(lambdan**2 - 15.3 * lambdan + 2050)
    if SectionType == 'CHS' or SectionType == 'RHS' or SectionType == 'SHS':
        alpha_b = -0.5
    elif SectionType == 'Universal_Beam' or SectionType == 'Universal_Column':
        if section_properties['tf'] <= 40:
            alpha_b = 0
    elif SectionType == 'PFC':
        alpha_b = 0.5
    else:
        alpha_b = 1
    lambda1 = lambdan + alpha_a * alpha_b
    n = max(0.00326 * (lambda1 -13.5),0)
    curly = ((lambda1/90)**2 + 1 + n)/(2*(lambda1/90)**2)
    alpha_c = curly * (1 - math.sqrt(1- (90/(curly*lambda1))**2))
    section_properties['Nc' + xy] = min(alpha_c*section_properties['Ns'+ xy],section_properties['Ns'+xy])
    section_properties['PhiNc'+ xy] = 0.9 * section_properties['Nc'+ xy]
    return section_properties

#This function will determine if the section is compact or not and then calculate the sectional moment capacity
def compact(section_properties, SectionType, L, alpha_m, restraint, load_height_position, longitudinal_position, ends_with_restraint):
    if SectionType == 'Universal_Beam' or SectionType == 'Universal_Column' or SectionType == 'PFC' or SectionType == 'T-section':
        if restraint == 'FF' or restraint == 'FL' or restraint == 'LL' or restraint == 'FU':
            kt = 1
        elif restraint == 'FP' or restraint == 'PL' or restraint == 'PU':
            kt = 1 + (((section_properties['d']-section_properties['tf']*2)/L/1000)*(section_properties['tf']/(2*section_properties['tw']))**3)
            print(kt,'THis is L')
        elif restraint == 'PP':
            kt = 1 + (2*((section_properties['d'] - section_properties['tf'] * 2) / L / 1000) * (
                        section_properties['tf'] / (2 * section_properties['tw'])) ** 3)
    elif SectionType == 'RHS'or SectionType == 'SHS':
        if restraint == 'FF' or restraint == 'FL' or restraint == 'LL' or restraint == 'FU':
            kt = 1
        elif restraint == 'FP' or restraint == 'PL' or restraint == 'PU':
            kt = 1 + (((section_properties['d'] - section_properties['tf'] * 2) / L / 1000) * (
                        section_properties['tf'] / (2 * section_properties['tw'])) ** 3)/2
        elif restraint == 'PP':
            kt = 1 + (2 * ((section_properties['d'] - section_properties['tf'] * 2) / L / 1000) * (
                    section_properties['tf'] / (2 * section_properties['tw'])) ** 3)/2
    Table_5_6_3 = wb['Table 5.6.3']
    count1 = 0
    for row in Table_5_6_3:
        count = 0
        if restraint in row[0].value and ends_with_restraint in row[1].value:
            row2 = count1
        if row[0].value == longitudinal_position and restraint in row[1].value:
            row1 = count1
        for cell in row:
            if load_height_position == cell.value:
                column = count
            count +=1
        count1 +=1
    kl = Table_5_6_3.cell(row = row1+1, column = column+1).value
    print(str(kl) + 'yes')
    kr = Table_5_6_3.cell(row = row2+1, column = 3).value
    if SectionType == 'Universal_Beam' or SectionType == 'Universal_Column' or SectionType == 'PFC' or SectionType == 'T-section' or SectionType == 'RHS':
        Le = L*kt*kl*kr
    else:
        Le =L
    print(Le)
    #Determine if section is compact from section properties
    #The 0.5 applies because the flange extends from each side of the web
    if SectionType == 'Universal_Beam' or SectionType == "Universal_Column" or SectionType == 'Welded_Beam' or SectionType == 'T-section':
        f_lambda_e = 0.5*(section_properties['bf'] / section_properties['tf'])*math.sqrt(section_properties['fyf'] / 250)
    elif SectionType == 'PFC':
        f_lambda_e = (section_properties['bf'] / section_properties['tf']) * math.sqrt(
            section_properties['fyf'] / 250)
    #Use Table 5.2 to find lambda values
    table_5_2 = wb['Table 5.2']
    if SectionType == 'Universal_Beam' or SectionType == "Universal_Column" or SectionType == 'Welded_Beam' or SectionType == 'PFC':
        for row in table_5_2:
            if row[0].value == 'UCOneSR':
                f_lambda_ey = row[6].value
                f_lambda_ep = row[5].value
                break
        for row in table_5_2:
            if row[0].value == 'CTOneSR':
                f_lambda_ey_OoP = row[6].value
                f_lambda_ep_OoP = row[5].value
                break
    else:
        f_lambda_e = 2
        f_lambda_ey = 1
        f_lambda_ep =1
        f_lambda_ey_OoP = 1
        f_lambda_ep_OoP = 1
    if SectionType == 'T-section':
        w_lambda_e = ((section_properties['d'] - section_properties['tf']) / section_properties['tw']) * math.sqrt(
            section_properties['fyw'] / 250)
    elif SectionType != 'T-section' and SectionType != 'CHS':
        w_lambda_e = ((section_properties['d']-2*section_properties['tf'])/ section_properties['tw']) * math.sqrt(
        section_properties['fyw'] / 250)
    # Use Table 5.2 to find lambda values
    table_5_2 = wb['Table 5.2']
    if SectionType != 'CHS':
        for row in table_5_2:
            if row[0].value == 'CTBothSR':
                w_lambda_ey = row[6].value
                w_lambda_ep = row[5].value
                break
        if f_lambda_e / f_lambda_ey > w_lambda_e / w_lambda_ey:
            lambda_s = f_lambda_e
            lambda_sy = f_lambda_ey
            lambda_sp = f_lambda_ep
        else:
            lambda_s = w_lambda_e
            lambda_sy = w_lambda_ey
            lambda_sp = w_lambda_ep
    if SectionType != 'RHS' and SectionType != 'SHS' and SectionType != 'CHS':
        if lambda_s < lambda_sp:
            section_properties['Zex'] = min(1.5*section_properties['Zx'],section_properties['Sx'])
            section_properties['compactness'] = 'compact'
            print(section_properties['Zex'])
        elif lambda_s > lambda_sp and lambda_s < lambda_sy:
            section_properties['Zex'] = section_properties['Zx'] + (((lambda_sy - lambda_s) / (lambda_sy - lambda_sp))*(min(1.5*section_properties['Zx'],section_properties['Sx']) - section_properties['Zx']))
            section_properties['compactness'] = 'non-compact'
        elif lambda_s > lambda_sy:
            section_properties['Zex'] = section_properties['Zx']*(lambda_sy/lambda_s)
            section_properties['compactness'] = 'slender'
        if f_lambda_e < f_lambda_ep_OoP:
            section_properties['Zey'] = min(1.5*section_properties['Zy'],section_properties['Sy'])
            section_properties['compactness OoP'] = 'compact'
            print(section_properties['Zey'])
        elif f_lambda_e > f_lambda_ep_OoP and f_lambda_e < f_lambda_ey_OoP:
            section_properties['Zey'] = section_properties['Zy'] + (((f_lambda_ey_OoP - f_lambda_e) / (f_lambda_ey_OoP - f_lambda_ep_OoP))*(min(1.5*section_properties['Zy'],section_properties['Sy']) - section_properties['Zy']))
            section_properties['compactness OoP'] = 'non-compact'
        elif f_lambda_e > f_lambda_ey_OoP:
            section_properties['Zey'] = section_properties['Zy']*(f_lambda_ey_OoP/f_lambda_e)
            section_properties['compactness OoP'] = 'slender'
    if SectionType != 'CHS':
        section_properties['Msx'] = min(section_properties['fyf'],section_properties['fyw'])*section_properties['Zex']/1000000
        section_properties['Msy'] = min(section_properties['fyf'],section_properties['fyw'])*section_properties['Zey']/1000000
        section_properties['PhiMsx'] = 0.9*section_properties['Msx']
        section_properties['PhiMsy'] = 0.9 * section_properties['Msy']
    else:
        section_properties['Msx'] = section_properties['fy'] * section_properties[
            'Ze'] / 1000000
        section_properties['Msy'] = section_properties['fy'] * section_properties[
            'Ze'] / 1000000
        section_properties['PhiMsx'] = 0.9 * section_properties['Msx']
        section_properties['PhiMsy'] = 0.9 * section_properties['Msy']
    if SectionType == 'T-section':
        Icy = section_properties['tf']*section_properties['tw']**3/12
        beta_x = 0.8*(section_properties['d'] - section_properties['tf'])*((2*Icy/section_properties['Iy'])-1)
        Moa = math.sqrt(math.pi**2*200*10**9*section_properties['Iy']*10**-12/(Le**2))*(math.sqrt(80*10**9*section_properties['J'] + (math.pi**2*200*10**-9*section_properties['Iw'])/(Le**2)
                                                                                                  + (beta_x**2*math.pi**2*200*10**9*section_properties['Iy'])/(4*Le**2)))  \
              + beta_x/2*math.sqrt(math.pi**2*200*10**9*section_properties['Iy']/(Le**2))
    elif SectionType == 'CHS':
        x =1
    else:
        Moa = math.sqrt(((math.pi**2 * 200 *10**9* section_properties['Iy']*10**(-12))/(Le**2))*(80*10**9*section_properties['J']*10**(-12) +((math.pi**2 * 200*10**9 * section_properties['Iw']*10**(-18))/(Le**2))))/1000
        Moa_OoP = math.sqrt(((math.pi ** 2 * 200 * 10 ** 9 * section_properties['Ix'] * 10 ** (-12)) / (Le ** 2)) * (
                    80 * 10 ** 9 * section_properties['J'] * 10 ** (-12) + (
                        (math.pi ** 2 * 200 * 10 ** 9 * section_properties['Iw'] * 10 ** (-18)) / (Le ** 2)))) / 1000
    if SectionType != 'CHS':
        alpha_s = 0.6*(math.sqrt((section_properties['Msx']/Moa)**2+3)-(section_properties['Msx']/Moa))
        alpha_s_OoP = 0.6 * (math.sqrt((section_properties['Msy'] / Moa_OoP) ** 2 + 3) - (section_properties['Msy'] / Moa_OoP))

        section_properties['PhiMbx'] = min(0.9*alpha_m*alpha_s*section_properties['Msx'],section_properties['PhiMsx'])
        section_properties['PhiMby'] = min(0.9 * alpha_m * alpha_s_OoP * section_properties['Msy'],
                                           section_properties['PhiMsy'])
        print(section_properties)
    #print(Moa)
        section_properties['Moa'] = Moa
        section_properties['alpha_m'] = alpha_m
        section_properties['alpha_s'] = alpha_s
    else:
        section_properties['PhiMbx'] = section_properties['PhiMsx']
        section_properties['PhiMby'] = section_properties['PhiMsy']
    #print(alpha_s)
    return section_properties

def T_section():
    x = 0

def shear(section_properties, SectionType):
    if SectionType == 'CHS':
        section_properties['Vw'] = 0.36*section_properties['fy'] * section_properties['Ag']
        section_properties['Vu'] = section_properties['Vw']
    elif SectionType != 'CHS':
        section_properties['Vw'] = 0.6 * section_properties['fyw'] * (section_properties['d'] - section_properties['tf']*2) * section_properties['tw'] / 1000
        if (section_properties['d'] - section_properties['tf'])/section_properties['tw'] <= 82/math.sqrt(section_properties['fyw']):
            section_properties['Vu'] = section_properties['Vw']
        elif (section_properties['d'] - section_properties['tf'])/section_properties['tw'] >= 82/math.sqrt(section_properties['fyw']):
            section_properties['alpha_v'] = (82/(((section_properties['d']-section_properties['tf']*2)/section_properties['tw'])*math.sqrt(section_properties['fyw']/250)))**2
            section_properties['Vu'] = min(section_properties['Vw']*section_properties['alpha_v'],section_properties['Vw'])
        if SectionType == 'SHS' or SectionType =='RHS':
            section_properties['Vu'] = 2*section_properties['Vu']
    section_properties['PhiVu'] = 0.9 * section_properties['Vu']
    return section_properties

def shear_moment(section_properties,M):
    if M <= 0.75*section_properties['PhiMsx']:
        section_properties['Vvm'] = section_properties['Vu']
        section_properties['PhiVvm'] = 0.9*section_properties['Vvm']
    elif M > 0.75*section_properties['PhiMsx'] and M < section_properties['PhiMsx']:
        section_properties['Vvm'] = section_properties['Vu']*(2.2-((1.6*M)/section_properties['PhiMsx']))
        section_properties['PhiVvm'] = 0.9 * section_properties['Vvm']
    else:
        section_properties['Vvm'] = 'M* exceeds PhiMsx'
    return section_properties


def deflection_check(section_properties,G_UDL,Q_UDL,G_PL,Q_PL,Length,OoP,W_UDL):
    E = 200
    load = G_UDL + 0.7*Q_UDL + W_UDL
    PL = G_PL + 0.7*Q_PL
    if OoP == True:
        I = section_properties['Iy']
    elif OoP == False:
        I = section_properties['Ix']
    UDL_deflection = (5*load*10**3*(Length)**4)/(384*E*I*10**-12*10**9)*10**3

    Point_load_deflection = PL*(Length)**3/(48*E*10**9*I*10**-12)*10**6
    load = 1.2*G_UDL + 1.5*Q_UDL + W_UDL
    PL = 1.2*G_PL + 1.5*Q_PL
    UDL_moment = load*(Length)**2/8
    UDL_shear = load*(Length)/2
    PL_moment = PL*(Length)/2
    PL_Shear = PL/2
    print(Point_load_deflection + UDL_deflection)
    return UDL_deflection, Point_load_deflection, UDL_moment,UDL_shear, PL_moment,PL_Shear

def printcalcs(SectionSize,section_properties,print_location,job_name,comments,M,V,N,Design_loads):
    geometry_options = {
        "head": "40pt",
        "margin": "0.5in",
        "bottom": "0.6in",
        "includeheadfoot": True
    }
    doc = Document(SectionSize,geometry_options=geometry_options)

    # Generating first page style
    first_page = PageStyle("firstpage")

    # Header image
    with first_page.create(Head("R")) as header_right:
        with header_right.create(MiniPage(width=NoEscape(r"0.49\textwidth"),
                                         pos='c')) as logo_wrapper:

            logo_wrapper.append(StandAloneGraphic(image_options="width=120px",
                                filename='Northrop_logo.png'))

    # Add docuement title
    with first_page.create(Head("R")) as right_header:
        with right_header.create(MiniPage(width=NoEscape(r"0.49\textwidth"),
                                 pos='c', align='r')) as title_wrapper:
            title_wrapper.append(LargeText(bold("Bank Account Statement")))
            title_wrapper.append(LineBreak())
            title_wrapper.append(MediumText(bold("Date")))

    # Add footer
    with first_page.create(Foot("C")) as footer:
        message = "Important message please read"
        with footer.create(Tabularx(
                "X X X X",
                width_argument=NoEscape(r"\textwidth"))) as footer_table:

            footer_table.add_row(
                [MultiColumn(4, align='l', data=TextColor("blue", message))])
            footer_table.add_hline(color="blue")
            footer_table.add_empty_row()

            branch_address = MiniPage(
                width=NoEscape(r"0.25\textwidth"),
                pos='t')
            branch_address.append("960 - 22nd street east")
            branch_address.append("\n")
            branch_address.append("Saskatoon, SK")

            docuement_details = MiniPage(width=NoEscape(r"0.25\textwidth"),
                                        pos='t', align='r')
            docuement_details.append("1000")
            docuement_details.append(LineBreak())
            docuement_details.append(simple_page_number())

            footer_table.add_row([branch_address, branch_address,
                                  branch_address, docuement_details])

    doc.preamble.append(first_page)
    # End first page style


    doc.change_document_style("firstpage")
    doc.add_color(name="lightgray", model="gray", description="0.80")




    with doc.create(Section(job_name)):
        doc.append(comments)
    if Design_loads == True:
        with doc.create(Subsection(SectionSize)):
            #Allow for inserting image at later stage
            if M == None:
                doc.append(NoEscape('$M^{*}$  = ' + M + ' KNm'))
                doc.append('\n\n')
            elif V == None:
                doc.append(NoEscape('$V^{*}$  = ' + M + ' KN'))
                doc.append('\n\n')
            elif N == None:
                doc.append(NoEscape('$N^{*}$  = ' + M + ' KN'))
                doc.append('\n\n')
    with doc.create(Section('TRY ' + SectionSize,numbering=False)):
        with doc.create(Tabular('lll')) as table:
            table.add_row('$\phi M_{sx}$ =  ', section_properties['PhiMsx'],' KNm',escape=False)
            table.add_empty_row()
            table.add_row('$\phi M_{bx}$ =  ', section_properties['PhiMbx'], ' KNm',escape=False)
            table.add_empty_row()
            table.add_row('$\phi M_{sy}$ =  ', section_properties['PhiMsy'], ' KNm',escape=False)
            table.add_empty_row()
            table.add_row('$\phi M_{by}$ =  ', section_properties['PhiMby'], ' KNm',escape=False)
            table.add_empty_row()
            #table.add_row('$\phi V_{u}$ =  ', section_properties['Vu'], ' KNm')
            table.add_empty_row()
            #table.add_row('$\phi N_{n}$ =  ', 'tbc', ' KNm')
            table.add_empty_row()





    doc.generate_tex()
    subprocess.run(['pdflatex','-interaction=nonstopmode', SectionSize+'.tex','-output-directory='+print_location, '-jobname='+job_name])

wb.close()